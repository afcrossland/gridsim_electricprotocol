#!/usr/bin/env python3
"""One-time import of the Electric Protocol policy map spreadsheet into seed JSON.

Reads "Electric Protocol - Policy Map - v2.xlsx" and emits src/data/protocol.seed.json.

The spreadsheet layout (see repo README):
  col B  question / heading text
  col C  scoring rubric, newline separated, "<score> = <description>"
  col D  impact weighting
  row 8  country columns from F onwards
  orange fill in col B marks a section heading

Two known defects in the source sheet are corrected here rather than in the
spreadsheet itself; both are reported on stdout when they are applied:

  1. Rows 26-33 have their first rubric tier mislabelled with an incrementing
     counter (1=, 2=, ... 8=) instead of 0=, apparently from a fill-down.
     Every one of these rubrics is renumbered to 0/1/2.
  2. Row 44 duplicates row 24 word for word at a different weight. Row 44 is
     dropped; row 24 is authoritative.
"""

import json
import re
import sys
from pathlib import Path

import openpyxl

REPO = Path(__file__).resolve().parent.parent
# The spreadsheet is committed under source/ so the import is reproducible from
# a clean checkout. The parent directory is kept as a fallback for the original
# working layout, where it sat alongside the app folder.
XLSX_NAME = "Electric Protocol - Policy Map - v2.xlsx"
XLSX = next(
    (p for p in (REPO / "source" / XLSX_NAME, REPO.parent / XLSX_NAME) if p.exists()),
    REPO / "source" / XLSX_NAME,
)
OUT = REPO / "src" / "data" / "protocol.seed.json"

FIRST_ROW, LAST_ROW = 9, 58  # row 9 is the first section heading
COUNTRY_ROW = 8
FIRST_COUNTRY_COL = 6  # F
HEADING_FILL = "FFF9B27E"  # orange

DROP_ROWS = {44}  # duplicate of row 24

# Share of total question weight a jurisdiction must answer before it is ranked
# and coloured on the map. Adjustable in the UI; this is the starting value.
COMPLETENESS_THRESHOLD = 0.3

# ISO 3166-1 alpha-2 for the sheet's country column labels. These keys are what
# the map colours by (world.geojson property `country_iso2`).
COUNTRY_ISO = {
    "NZ": "NZ",
    "GB": "GB",
    "India": "IN",
    "Sri Lanka": "LK",
    "Pakistan": "PK",
    "Malaysia": "MY",
}
COUNTRY_NAME = {
    "NZ": "New Zealand",
    "GB": "Great Britain",
    "India": "India",
    "Sri Lanka": "Sri Lanka",
    "Pakistan": "Pakistan",
    "Malaysia": "Malaysia",
}

# Separator is normally "=", but at least one row (10) uses "-" instead.
TIER_RE = re.compile(r"^\s*(\d+)\s*[=-]\s*(.*)$", re.S)

notes = []


def clean_section_title(text):
    """Strip the redundant "Electric Protocol" prefix from a section heading.

    Seven of the nine headings are written as "Electric Protocol - <name>" in
    the spreadsheet, using both hyphens and en dashes. Inside an application
    already called Electric Protocol the prefix says nothing, and it pushes the
    part that identifies the section out of a narrow rail.
    """
    return re.sub(r"^\s*Electric Protocol\s*[-–—]\s*", "", text).strip()


def slug(text, maxlen=48):
    s = re.sub(r"[^a-z0-9]+", "-", text.lower()).strip("-")
    return s[:maxlen].rstrip("-")


def parse_rubric(raw, row):
    """Parse a rubric cell into ordered {score, label} tiers."""
    tiers = []
    for line in str(raw).split("\n"):
        if not line.strip():
            continue
        m = TIER_RE.match(line)
        if not m:
            notes.append(f"row {row}: rubric line without a score prefix: {line.strip()[:60]!r}")
            continue
        tiers.append({"score": int(m.group(1)), "label": m.group(2).strip()})

    if not tiers:
        raise ValueError(f"row {row}: no rubric tiers parsed")

    # Defect 1: first tier should always be the zero case. When it is not, the
    # whole rubric is renumbered positionally.
    if tiers[0]["score"] != 0:
        before = [t["score"] for t in tiers]
        if len(tiers) == 3:
            for i, t in enumerate(tiers):
                t["score"] = i
        elif len(tiers) == 2:
            tiers[0]["score"], tiers[1]["score"] = 0, 2
        else:
            raise ValueError(f"row {row}: cannot renumber a {len(tiers)}-tier rubric")
        notes.append(
            f"row {row}: renumbered corrupt rubric {before} -> {[t['score'] for t in tiers]}"
        )

    scores = [t["score"] for t in tiers]
    if scores != sorted(scores) or scores[0] != 0 or scores[-1] != 2:
        raise ValueError(f"row {row}: rubric scores {scores} are not a valid 0..2 ladder")
    return tiers


def main():
    if not XLSX.exists():
        sys.exit(f"spreadsheet not found: {XLSX}")

    ws = openpyxl.load_workbook(XLSX, data_only=True).active

    # Country columns
    countries = []
    for col in range(FIRST_COUNTRY_COL, ws.max_column + 1):
        label = ws.cell(COUNTRY_ROW, col).value
        if not label:
            continue
        label = str(label).strip()
        countries.append(
            {
                "col": col,
                "code": COUNTRY_ISO.get(label, label),
                "name": COUNTRY_NAME.get(label, label),
            }
        )

    sections = []
    questions = []
    current_section = None
    current_subsection = None

    for row in range(FIRST_ROW, LAST_ROW + 1):
        text = ws.cell(row, 2).value
        if not text:
            continue
        text = str(text).strip()
        rubric_cell = ws.cell(row, 3).value
        fill = ws.cell(row, 2).fill
        is_heading = bool(fill and fill.patternType and fill.fgColor.rgb == HEADING_FILL)

        if is_heading:
            title = clean_section_title(text)
            current_section = {
                "id": slug(title),
                "title": title,
                "sourceRow": row,
                "order": len(sections),
            }
            sections.append(current_section)
            current_subsection = None
            continue

        if rubric_cell is None:
            # Unfilled heading row: a subsection label grouping the rows below it.
            current_subsection = text
            continue

        if row in DROP_ROWS:
            notes.append(f"row {row}: dropped as a duplicate ({text[:50]!r})")
            continue

        if current_section is None:
            raise ValueError(f"row {row}: question before any section heading")

        weight = ws.cell(row, 4).value
        if weight is None:
            raise ValueError(f"row {row}: missing weight")

        answers = {}
        for c in countries:
            v = ws.cell(row, c["col"]).value
            if v is None:
                continue
            answers[c["code"]] = int(v)

        questions.append(
            {
                "id": f"q{row:03d}",
                "sourceRow": row,
                "sectionId": current_section["id"],
                "subsection": current_subsection,
                "order": len(questions),
                "text": text,
                "weight": float(weight),
                "rubric": parse_rubric(rubric_cell, row),
                "seedAnswers": answers,
            }
        )

    # Preamble definitions (rows 4-5) carried through as glossary copy.
    glossary = [
        str(ws.cell(r, 2).value).strip()
        for r in (4, 5)
        if ws.cell(r, 2).value
    ]

    seed = {
        "title": str(ws.cell(2, 2).value or "Electric Protocol").strip(),
        "source": XLSX.name,
        "maxScore": 2,
        "completenessThreshold": COMPLETENESS_THRESHOLD,
        "glossary": glossary,
        "sections": [
            {k: v for k, v in s.items() if k != "col"} for s in sections
        ],
        "countries": [
            {"code": c["code"], "name": c["name"]}
            for c in countries
        ],
        "questions": questions,
    }

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(seed, indent=2, ensure_ascii=False) + "\n")

    answered = {c["code"]: 0 for c in countries}
    for q in questions:
        for code in q["seedAnswers"]:
            answered[code] += 1

    print(f"wrote {OUT.relative_to(REPO)}")
    print(f"  {len(sections)} sections, {len(questions)} questions, "
          f"total weight {sum(q['weight'] for q in questions):g}")
    for code, n in answered.items():
        print(f"  {code}: {n}/{len(questions)} answered")
    if notes:
        print("corrections applied:")
        for n in notes:
            print(f"  - {n}")


if __name__ == "__main__":
    main()
